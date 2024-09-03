import openpyxl
workbook = openpyxl.load_workbook('test.xlsx')
sheet = workbook.active

# 遍历单元格比较数据
row_num = 1
while row_num < sheet.max_row:
    if sheet.cell(row=row_num, column=1).value == sheet.cell(row=row_num+1, column=1).value and sheet.cell(row=row_num, column=2).value == sheet.cell(row=row_num+1, column=2).value:
        sheet.cell(row=row_num, column=3, value=1)
        sheet.cell(row=row_num+1, column=3, value=1)
    row_num += 1

# 保存修改后的 Excel 文件
workbook.save('test1.xlsx')